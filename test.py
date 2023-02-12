import json
import os
from pathlib import Path

import openpyxl

from main import (build_excel_file_of_imgur_album_links, create_album,
                  delete_all_albums, iterate_through_vids, upload_mp4_to_imgur,
                  upload_video_to_album, write_csv_to_excel,
                  write_string_list_to_csv)

def test_create_album():
    ALBUM_NAME = "Test album"
    ALBUM_DESCRIPTION = "Test album description"
    response = create_album(ALBUM_NAME, ALBUM_DESCRIPTION)
    assert response.status_code == 200
    json_response = json.loads(response.text)
    assert json_response.get("success") is True
    assert json_response.get("status") == 200
    assert json_response.get("data").get("id") is not None
    assert json_response.get("data").get("deletehash") is not None


def test_iterate_through_vids():
    count = iterate_through_vids("samples", print)
    video_list = os.listdir("samples")
    assert len(video_list) == count


def test_upload_to_imgur():
    VIDEO_FILENAME = "samples/file_example_AVI_480_750kB.avi"
    response = upload_mp4_to_imgur(VIDEO_FILENAME)
    assert response.status_code == 200
    json_response = json.loads(response.text)
    assert json_response.get("success") is True
    assert json_response.get("status") == 200
    assert json_response.get("data").get("id") is not None
    assert json_response.get("data").get("deletehash") is not None
    assert json_response.get("data").get("link") is not None


def test_upload_to_album():
    ALBUM_NAME = "Test an album with a video"
    ALBUM_DESCRIPTION = "Test album description"
    VIDEO_FILENAME = "samples/file_example_AVI_480_750kB.avi"
    link_to_video_in_album = upload_video_to_album(ALBUM_NAME, ALBUM_DESCRIPTION, VIDEO_FILENAME)
    assert '/a' in link_to_video_in_album


def test_add_to_excel():
    strings = ['test1', 'test2', 'test3']
    input_file_name = 'test_file.csv'
    write_string_list_to_csv(strings, input_file_name)
    excel_filename = write_csv_to_excel(input_file_name)
    assert Path(excel_filename).is_file()
    assert openpyxl.load_workbook(excel_filename).active.max_row == len(strings)
    assert openpyxl.load_workbook(excel_filename).active['A1'].value == strings[0]
    os.remove(input_file_name)
    os.remove(excel_filename)


def test_end_to_end_excel_file():
    delete_all_albums()
    excel_filename = build_excel_file_of_imgur_album_links("samples", "samples_test")
    video_count = iterate_through_vids("samples", print)
    assert Path(excel_filename).is_file()
    assert openpyxl.load_workbook(excel_filename).active.max_row == video_count
    assert openpyxl.load_workbook(excel_filename).active['A1'].value is not None
    assert '/a/' in openpyxl.load_workbook(excel_filename).active['A1'].value
